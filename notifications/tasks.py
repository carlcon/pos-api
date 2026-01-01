"""
Celery tasks for notifications app.
"""
from celery import shared_task
from django.utils import timezone
from django.conf import settings
import os
import csv
import logging

logger = logging.getLogger(__name__)


@shared_task(bind=True)
def process_export_job(self, job_id):
    """
    Process an export job in the background.
    """
    from notifications.models import ExportJob, Notification
    
    try:
        job = ExportJob.objects.get(id=job_id)
    except ExportJob.DoesNotExist:
        logger.error(f"Export job {job_id} not found")
        return
    
    job.status = ExportJob.Status.PROCESSING
    job.save(update_fields=['status'])
    
    try:
        # Create exports directory if it doesn't exist
        export_dir = settings.EXPORT_FILES_DIR
        os.makedirs(export_dir, exist_ok=True)
        
        # Process based on export type
        if job.export_type == ExportJob.ExportType.SALES_CSV:
            file_path = export_sales_csv(job, export_dir)
        elif job.export_type == ExportJob.ExportType.SALES_EXCEL:
            file_path = export_sales_excel(job, export_dir)
        elif job.export_type == ExportJob.ExportType.SALES_PDF:
            file_path = export_sales_pdf(job, export_dir)
        elif job.export_type == ExportJob.ExportType.PRODUCTS_CSV:
            file_path = export_products_csv(job, export_dir)
        elif job.export_type == ExportJob.ExportType.PRODUCTS_EXCEL:
            file_path = export_products_excel(job, export_dir)
        elif job.export_type == ExportJob.ExportType.STOCK_CSV:
            file_path = export_stock_csv(job, export_dir)
        elif job.export_type == ExportJob.ExportType.STOCK_EXCEL:
            file_path = export_stock_excel(job, export_dir)
        else:
            raise ValueError(f"Unknown export type: {job.export_type}")
        
        job.file_path = file_path
        job.status = ExportJob.Status.COMPLETED
        job.progress = 100
        job.completed_at = timezone.now()
        job.save()
        
        # Create notification
        Notification.objects.create(
            user=job.user,
            type=Notification.Type.EXPORT_COMPLETE,
            title='Export Complete',
            message=f'Your {job.get_export_type_display()} export is ready for download.',
            data={'job_id': job.id, 'export_type': job.export_type}
        )
        
        logger.info(f"Export job {job_id} completed successfully")
        
    except Exception as e:
        logger.error(f"Export job {job_id} failed: {e}")
        job.status = ExportJob.Status.FAILED
        job.error_message = str(e)
        job.save()


def export_sales_csv(job, export_dir):
    """Export sales to CSV."""
    from sales.models import Sale
    
    filters = job.filters or {}
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f'sales_export_{timestamp}.csv'
    file_path = os.path.join(export_dir, filename)
    
    # Build queryset
    queryset = Sale.objects.select_related('cashier', 'store').prefetch_related('items')
    
    if filters.get('store_id'):
        queryset = queryset.filter(store_id=filters['store_id'])
    if filters.get('date_from'):
        queryset = queryset.filter(created_at__gte=filters['date_from'])
    if filters.get('date_to'):
        queryset = queryset.filter(created_at__lte=filters['date_to'])
    if filters.get('payment_method'):
        queryset = queryset.filter(payment_method=filters['payment_method'])
    if filters.get('partner_id'):
        queryset = queryset.filter(partner_id=filters['partner_id'])
    
    # Write CSV
    with open(file_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            'Sale Number', 'Date', 'Customer', 'Cashier', 'Store',
            'Payment Method', 'Subtotal', 'Discount', 'Total', 'Items Count'
        ])
        
        total_count = queryset.count()
        for i, sale in enumerate(queryset.iterator()):
            writer.writerow([
                sale.sale_number,
                sale.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                sale.customer_name or '',
                sale.cashier.username if sale.cashier else '',
                sale.store.name if sale.store else '',
                sale.get_payment_method_display(),
                sale.subtotal,
                sale.discount,
                sale.total_amount,
                sale.items.count()
            ])
            
            # Update progress
            if i % 100 == 0:
                job.progress = int((i / total_count) * 90)
                job.save(update_fields=['progress'])
    
    return file_path


def export_sales_excel(job, export_dir):
    """Export sales to Excel."""
    from sales.models import Sale
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    filters = job.filters or {}
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f'sales_export_{timestamp}.xlsx'
    file_path = os.path.join(export_dir, filename)
    
    # Build queryset
    queryset = Sale.objects.select_related('cashier', 'store').prefetch_related('items')
    
    if filters.get('store_id'):
        queryset = queryset.filter(store_id=filters['store_id'])
    if filters.get('date_from'):
        queryset = queryset.filter(created_at__gte=filters['date_from'])
    if filters.get('date_to'):
        queryset = queryset.filter(created_at__lte=filters['date_to'])
    if filters.get('payment_method'):
        queryset = queryset.filter(payment_method=filters['payment_method'])
    if filters.get('partner_id'):
        queryset = queryset.filter(partner_id=filters['partner_id'])
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    
    # Header styling
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='049AE0', end_color='049AE0', fill_type='solid')
    
    headers = [
        'Sale Number', 'Date', 'Customer', 'Cashier', 'Store',
        'Payment Method', 'Subtotal', 'Discount', 'Total', 'Items Count'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    total_count = queryset.count()
    for i, sale in enumerate(queryset.iterator(), 2):
        ws.cell(row=i, column=1, value=sale.sale_number)
        ws.cell(row=i, column=2, value=sale.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=i, column=3, value=sale.customer_name or '')
        ws.cell(row=i, column=4, value=sale.cashier.username if sale.cashier else '')
        ws.cell(row=i, column=5, value=sale.store.name if sale.store else '')
        ws.cell(row=i, column=6, value=sale.get_payment_method_display())
        ws.cell(row=i, column=7, value=float(sale.subtotal))
        ws.cell(row=i, column=8, value=float(sale.discount))
        ws.cell(row=i, column=9, value=float(sale.total_amount))
        ws.cell(row=i, column=10, value=sale.items.count())
        
        # Update progress
        if i % 100 == 0:
            job.progress = int((i / total_count) * 90)
            job.save(update_fields=['progress'])
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    wb.save(file_path)
    return file_path


def export_sales_pdf(job, export_dir):
    """Export sales to PDF."""
    from sales.models import Sale
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    
    filters = job.filters or {}
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f'sales_export_{timestamp}.pdf'
    file_path = os.path.join(export_dir, filename)
    
    # Build queryset
    queryset = Sale.objects.select_related('cashier', 'store').prefetch_related('items')
    
    if filters.get('store_id'):
        queryset = queryset.filter(store_id=filters['store_id'])
    if filters.get('date_from'):
        queryset = queryset.filter(created_at__gte=filters['date_from'])
    if filters.get('date_to'):
        queryset = queryset.filter(created_at__lte=filters['date_to'])
    if filters.get('payment_method'):
        queryset = queryset.filter(payment_method=filters['payment_method'])
    if filters.get('partner_id'):
        queryset = queryset.filter(partner_id=filters['partner_id'])
    
    # Create PDF
    doc = SimpleDocTemplate(file_path, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center
    )
    elements.append(Paragraph('Sales Report', title_style))
    elements.append(Spacer(1, 0.25*inch))
    
    # Table data
    data = [[
        'Sale #', 'Date', 'Customer', 'Cashier', 'Store',
        'Payment', 'Total'
    ]]
    
    for sale in queryset[:500]:  # Limit to 500 for PDF
        data.append([
            sale.sale_number,
            sale.created_at.strftime('%Y-%m-%d'),
            (sale.customer_name or '')[:20],
            (sale.cashier.username if sale.cashier else '')[:15],
            (sale.store.name if sale.store else '')[:15],
            sale.get_payment_method_display()[:10],
            f'{sale.total_amount:,.2f}'
        ])
    
    # Create table
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#049AE0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    return file_path


def export_products_csv(job, export_dir):
    """Export products to CSV."""
    from inventory.models import Product
    
    filters = job.filters or {}
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f'products_export_{timestamp}.csv'
    file_path = os.path.join(export_dir, filename)
    
    queryset = Product.objects.select_related('category', 'store')
    
    if filters.get('store_id'):
        queryset = queryset.filter(store_id=filters['store_id'])
    if filters.get('partner_id'):
        queryset = queryset.filter(partner_id=filters['partner_id'])
    if filters.get('category_id'):
        queryset = queryset.filter(category_id=filters['category_id'])
    
    with open(file_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            'Name', 'SKU', 'Barcode', 'Category', 'Store', 'Quantity',
            'Retail Price', 'Wholesale Price', 'Cost Price', 'Status'
        ])
        
        for product in queryset.iterator():
            writer.writerow([
                product.name,
                product.sku,
                product.barcode,
                product.category.name if product.category else '',
                product.store.name if product.store else '',
                product.quantity,
                product.retail_price,
                product.wholesale_price,
                product.cost_price,
                'Active' if product.is_active else 'Inactive'
            ])
    
    return file_path


def export_products_excel(job, export_dir):
    """Export products to Excel."""
    from inventory.models import Product
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    filters = job.filters or {}
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f'products_export_{timestamp}.xlsx'
    file_path = os.path.join(export_dir, filename)
    
    queryset = Product.objects.select_related('category', 'store')
    
    if filters.get('store_id'):
        queryset = queryset.filter(store_id=filters['store_id'])
    if filters.get('partner_id'):
        queryset = queryset.filter(partner_id=filters['partner_id'])
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='049AE0', end_color='049AE0', fill_type='solid')
    
    headers = [
        'Name', 'SKU', 'Barcode', 'Category', 'Store', 'Quantity',
        'Retail Price', 'Wholesale Price', 'Cost Price', 'Status'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    for i, product in enumerate(queryset.iterator(), 2):
        ws.cell(row=i, column=1, value=product.name)
        ws.cell(row=i, column=2, value=product.sku)
        ws.cell(row=i, column=3, value=product.barcode)
        ws.cell(row=i, column=4, value=product.category.name if product.category else '')
        ws.cell(row=i, column=5, value=product.store.name if product.store else '')
        ws.cell(row=i, column=6, value=product.quantity)
        ws.cell(row=i, column=7, value=float(product.retail_price))
        ws.cell(row=i, column=8, value=float(product.wholesale_price))
        ws.cell(row=i, column=9, value=float(product.cost_price))
        ws.cell(row=i, column=10, value='Active' if product.is_active else 'Inactive')
    
    wb.save(file_path)
    return file_path


def export_stock_csv(job, export_dir):
    """Export stock transactions to CSV."""
    from stock.models import StockTransaction
    
    filters = job.filters or {}
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f'stock_export_{timestamp}.csv'
    file_path = os.path.join(export_dir, filename)
    
    queryset = StockTransaction.objects.select_related('product', 'store', 'created_by')
    
    if filters.get('store_id'):
        queryset = queryset.filter(store_id=filters['store_id'])
    if filters.get('date_from'):
        queryset = queryset.filter(created_at__gte=filters['date_from'])
    if filters.get('date_to'):
        queryset = queryset.filter(created_at__lte=filters['date_to'])
    
    with open(file_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            'Date', 'Product', 'Store', 'Type', 'Quantity',
            'Before', 'After', 'User', 'Notes'
        ])
        
        for txn in queryset.iterator():
            writer.writerow([
                txn.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                txn.product.name if txn.product else '',
                txn.store.name if txn.store else '',
                txn.get_transaction_type_display(),
                txn.quantity,
                txn.quantity_before,
                txn.quantity_after,
                txn.created_by.username if txn.created_by else '',
                txn.notes or ''
            ])
    
    return file_path


def export_stock_excel(job, export_dir):
    """Export stock transactions to Excel."""
    from stock.models import StockTransaction
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    filters = job.filters or {}
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f'stock_export_{timestamp}.xlsx'
    file_path = os.path.join(export_dir, filename)
    
    queryset = StockTransaction.objects.select_related('product', 'store', 'created_by')
    
    if filters.get('store_id'):
        queryset = queryset.filter(store_id=filters['store_id'])
    if filters.get('date_from'):
        queryset = queryset.filter(created_at__gte=filters['date_from'])
    if filters.get('date_to'):
        queryset = queryset.filter(created_at__lte=filters['date_to'])
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Transactions"
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='049AE0', end_color='049AE0', fill_type='solid')
    
    headers = [
        'Date', 'Product', 'Store', 'Type', 'Quantity',
        'Before', 'After', 'User', 'Notes'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    for i, txn in enumerate(queryset.iterator(), 2):
        ws.cell(row=i, column=1, value=txn.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=i, column=2, value=txn.product.name if txn.product else '')
        ws.cell(row=i, column=3, value=txn.store.name if txn.store else '')
        ws.cell(row=i, column=4, value=txn.get_transaction_type_display())
        ws.cell(row=i, column=5, value=txn.quantity)
        ws.cell(row=i, column=6, value=txn.quantity_before)
        ws.cell(row=i, column=7, value=txn.quantity_after)
        ws.cell(row=i, column=8, value=txn.created_by.username if txn.created_by else '')
        ws.cell(row=i, column=9, value=txn.notes or '')
    
    wb.save(file_path)
    return file_path


@shared_task
def check_stock_levels():
    """
    Check stock levels and create notifications for low/out of stock items.
    Runs daily via Celery Beat.
    """
    from django.db.models import F
    from inventory.models import Product
    from notifications.utils import create_stock_alert_notifications
    
    logger.info("Running stock level check...")
    
    # Find low stock products
    low_stock_products = Product.objects.filter(
        is_active=True,
        quantity__gt=0,
        quantity__lte=F('low_stock_threshold')
    ).select_related('store')
    
    for product in low_stock_products:
        create_stock_alert_notifications(product, alert_type='low')
    
    # Find out of stock products
    out_of_stock_products = Product.objects.filter(
        is_active=True,
        quantity=0
    ).select_related('store')
    
    for product in out_of_stock_products:
        create_stock_alert_notifications(product, alert_type='out')
    
    logger.info(f"Stock check complete. Low: {low_stock_products.count()}, Out: {out_of_stock_products.count()}")


@shared_task
def cleanup_old_exports():
    """
    Clean up export files older than retention period.
    Runs daily via Celery Beat.
    """
    from notifications.models import ExportJob
    from datetime import timedelta
    
    retention_days = getattr(settings, 'EXPORT_FILE_RETENTION_DAYS', 1)
    cutoff_date = timezone.now() - timedelta(days=retention_days)
    
    old_jobs = ExportJob.objects.filter(
        created_at__lt=cutoff_date,
        status=ExportJob.Status.COMPLETED
    )
    
    deleted_count = 0
    for job in old_jobs:
        if job.file_path and os.path.exists(job.file_path):
            try:
                os.remove(job.file_path)
                deleted_count += 1
            except Exception as e:
                logger.error(f"Failed to delete export file {job.file_path}: {e}")
        job.delete()
    
    logger.info(f"Cleaned up {deleted_count} old export files")
